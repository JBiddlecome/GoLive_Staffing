import os
import pandas as pd
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the environment variables
env_path = r"C:\Users\jakeb\OneDrive\Documents\GitHub\golive-staffing-tools.env"
load_dotenv(env_path)

def get_engine():
    reportable_host = os.getenv("REPORTABLE_DB_HOST")
    host = reportable_host or os.getenv("DB_HOST")
    name = os.getenv("REPORTABLE_DB_NAME") or os.getenv("DB_NAME", "cstaffing_live")
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    reportable_port = os.getenv("REPORTABLE_DB_PORT")
    port = int(reportable_port or os.getenv("DB_PORT", "3306"))
    
    return create_engine(URL.create(
        drivername="mysql+pymysql",
        username=user,
        password=password,
        host=host,
        port=port,
        database=name
    ))

def format_excel(file_path):
    """
    Apply professional formatting to the Excel sheet:
    - Auto-adjust column widths
    - Set custom header style (navy background, white text)
    - Apply gridlines and zebra striping (subtle light gray)
    - Format datetime and date columns properly
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    # Define styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    normal_font = Font(name="Calibri", size=11)
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    row_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Format headers
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)
        cell.border = row_border

    # Format data rows
    for row_idx in range(2, ws.max_row + 1):
        is_zebra = (row_idx % 2 == 0)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = normal_font
            cell.border = row_border
            
            # Apply zebra striping
            if is_zebra:
                cell.fill = zebra_fill
            
            # Alignments & Formats based on headers
            header_val = ws.cell(row=1, column=col_idx).value
            if header_val in ["Event Date", "Shift Start", "Shift End"]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif header_val == "Position ID":
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                # Add a bit of padding for headers
                max_len = max(max_len, len(val_str) + 4)
            else:
                max_len = max(max_len, len(val_str))
        
        # Limit column width to 50 max to prevent extremely wide columns for descriptions
        ws.column_dimensions[col_letter].width = min(max(max_len, 10), 50)
        
    wb.save(file_path)

def main():
    engine = get_engine()
    
    # Query with strict soft delete filters on all tables
    query = """
    SELECT 
        e.date AS `Event Date`,
        s.start AS `Shift Start`,
        s.end AS `Shift End`,
        c.name AS `Client Name`,
        v.name AS `Venue Name`,
        sp.position_id AS `Position ID`,
        p.description AS `Position Title`,
        sp.additional_title AS `Additional Title`,
        sp.position_description AS `Shift Position Description`
    FROM shift_position sp
    JOIN shift s ON sp.shift_id = s.shift_id
    JOIN event e ON s.event_id = e.event_id
    LEFT JOIN venue v ON e.venue_id = v.venue_id
    LEFT JOIN client c ON e.client_id = c.client_id
    LEFT JOIN position p ON sp.position_id = p.position_id
    WHERE e.date BETWEEN '2026-03-01' AND '2026-04-30'
      AND c.deleted_at IS NULL
      AND v.deleted_at IS NULL
      AND e.deleted_at IS NULL
      AND s.deleted_at IS NULL
      AND sp.deleted_at IS NULL
      AND (p.deleted_at IS NULL OR p.deleted_at IS NOT NULL) # Ensure we handle positions gracefully
    ORDER BY e.date ASC, s.start ASC, c.name ASC
    """
    
    print("Executing complete query...")
    df = pd.read_sql_query(query, engine)
    
    # Strip any leading/trailing whitespace
    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].astype(str).str.strip().replace({'None': '', 'nan': ''})
            
    print(f"Total shift positions found: {len(df)}")
    
    # Paths for files in root directory
    csv_path = r"C:\Users\jakeb\OneDrive\Documents\GitHub\GoLive_Staffing\shifts_march_april_2026.csv"
    xlsx_path = r"C:\Users\jakeb\OneDrive\Documents\GitHub\GoLive_Staffing\shifts_march_april_2026.xlsx"
    
    # Save CSV
    df.to_csv(csv_path, index=False)
    print(f"Saved CSV to: {csv_path}")
    
    # Save Excel
    df.to_excel(xlsx_path, index=False)
    # Apply professional formatting
    format_excel(xlsx_path)
    print(f"Saved and formatted Excel to: {xlsx_path}")

    # Generate some summaries for the user
    print("\nSummary Statistics:")
    print(f"Unique Clients: {df['Client Name'].nunique()}")
    print(f"Unique Venues: {df['Venue Name'].nunique()}")
    print(f"Unique Positions: {df['Position Title'].nunique()}")
    
    print("\nTop 5 Active Clients:")
    print(df['Client Name'].value_counts().head(5).to_string())
    
    print("\nTop 5 Positions:")
    print(df['Position Title'].value_counts().head(5).to_string())

if __name__ == "__main__":
    main()

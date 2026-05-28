import os
import pandas as pd
import numpy as np
from sqlalchemy import create_engine, text
from sqlalchemy.engine import URL
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Load the environment variables
env_path = r"C:\Users\jakeb\OneDrive\Documents\GitHub\golive-staffing-tools.env"
load_dotenv(env_path)

def get_engine():
    reportable_host = os.getenv("REPORTABLE_DB_HOST")
    host = reportable_host or os.getenv("DB_HOST")
    name = os.getenv("REPORTABLE_DB_NAME") or os.getenv("DB_NAME", "cstaffing_live")
    user = os.getenv("DB_USER")
    password = os.getenv("DB_PASSWORD")
    reportable_port = os.getenv("REPORTABLE_DB_PORT")
    port = int(reportable_port or os.getenv("DB_PORT", "3306"))
    
    return create_engine(URL.create(
        drivername="mysql+pymysql",
        username=user,
        password=password,
        host=host,
        port=port,
        database=name
    ))

def format_excel(file_path):
    """
    Applies professional styling to the Excel report:
    - Deep navy headers with white text
    - Subtle zebra striping for rows
    - Grid lines enabled
    - Right-align numbers, center dates/IDs, left-align text
    - Auto-adjust column widths
    """
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.views.sheetView[0].showGridLines = True
    
    # Define color scheme (Corporate Navy Theme)
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    zebra_fill = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    normal_font = Font(name="Segoe UI", size=10)
    
    border_thin = Side(border_style="thin", color="D9D9D9")
    row_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
    
    # Header styling
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
        cell.border = row_border
    
    # Data rows styling
    for row_idx in range(2, ws.max_row + 1):
        is_zebra = (row_idx % 2 == 0)
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = normal_font
            cell.border = row_border
            
            if is_zebra:
                cell.fill = zebra_fill
                
            header_val = ws.cell(row=1, column=col_idx).value
            
            # Formatting based on columns
            if "Date" in header_val:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "ID" in header_val:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif "Seconds" in header_val:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "#,##0"
            elif "Hours" in header_val:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = "0.00"
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Set row heights
    ws.row_dimensions[1].height = 28
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20

    # Auto-fit columns with safety margin
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val_str = str(cell.value or '')
            if cell.row == 1:
                max_len = max(max_len, len(val_str) + 4)
            else:
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(min(max_len, 40), 12)
        
    wb.save(file_path)

def main():
    engine = get_engine()
    
    query = """
    SELECT 
        e.date AS `Event Date`,
        c.name AS `Client Name`,
        v.name AS `Venue Name`,
        p.description AS `Position Title`,
        vp.venue_position_id AS `Venue Position ID`,
        se.shift_employee_id AS `Shift Employee ID`,
        CONCAT(emp.first_name, ' ', emp.last_name) AS `Employee Name`,
        t.use_sheet AS `Use Sheet Setting`,
        COALESCE(t.employee_seconds, 0) AS `Employee Seconds`,
        COALESCE(t.client_seconds, 0) AS `Client Seconds`
    FROM timesheet t
    JOIN shift_employee se ON se.shift_employee_id = t.shift_employee_id
    JOIN shift_position sp ON sp.shift_position_id = se.shift_position_id
    JOIN shift s ON s.shift_id = sp.shift_id
    JOIN event e ON e.event_id = t.event_id
    JOIN client c ON c.client_id = e.client_id
    JOIN venue v ON v.venue_id = e.venue_id
    JOIN employee emp ON emp.employee_id = se.employee_id
    LEFT JOIN position p ON sp.position_id = p.position_id
    LEFT JOIN venue_position vp ON vp.venue_id = e.venue_id AND vp.position_id = sp.position_id
    WHERE e.date BETWEEN '2026-04-25' AND '2026-05-24'
      AND c.deleted_at IS NULL
      AND v.deleted_at IS NULL
      AND e.deleted_at IS NULL
      AND s.deleted_at IS NULL
      AND sp.deleted_at IS NULL
      AND se.deleted_at IS NULL
      AND emp.deleted_at IS NULL
    ORDER BY e.date ASC, c.name ASC, emp.last_name ASC, emp.first_name ASC
    """
    
    print("Executing query...")
    df = pd.read_sql_query(query, engine)
    print(f"Total shift assignments (timesheets) fetched: {len(df)}")
    
    if df.empty:
        print("No shift assignments found for the specified period.")
        return
        
    # Calculate Authoritative Seconds Worked
    # If use_sheet is CLIENT -> client_seconds
    # If use_sheet is EMPLOYEE -> employee_seconds
    # If use_sheet is NULL/empty -> default to employee_seconds for physical worked seconds,
    # but let's calculate worked seconds exactly matching business rules.
    def get_worked_seconds(row):
        use_sheet = row['Use Sheet Setting']
        if use_sheet == 'CLIENT':
            return row['Client Seconds']
        elif use_sheet == 'EMPLOYEE':
            return row['Employee Seconds']
        else:
            return row['Employee Seconds'] # Default pay/worked seconds
            
    df['Seconds Worked'] = df.apply(get_worked_seconds, axis=1)
    df['Hours Worked'] = df['Seconds Worked'] / 3600.0
    
    # Handle missing/NULL Venue Position IDs gracefully for excel representation
    df['Venue Position ID'] = df['Venue Position ID'].fillna('').replace({np.nan: ''})
    
    # Reorder columns to place Client Name, Seconds Worked, and Venue Position ID in prominent positions,
    # but keep other extremely useful columns for context.
    columns_ordered = [
        'Event Date',
        'Client Name',
        'Venue Position ID',
        'Seconds Worked',
        'Hours Worked',
        'Venue Name',
        'Position Title',
        'Employee Name',
        'Use Sheet Setting',
        'Employee Seconds',
        'Client Seconds',
        'Shift Employee ID'
    ]
    df_output = df[columns_ordered]
    
    output_path = r"C:\Users\jakeb\OneDrive\Documents\GitHub\GoLive_Staffing\shifts_report_april_may_2026.xlsx"
    print("Writing to Excel...")
    df_output.to_excel(output_path, index=False)
    
    print("Applying styling...")
    format_excel(output_path)
    print(f"Report generated successfully and saved to: {output_path}")
    
    # Show summary statistics
    print("\nReport Summary Statistics:")
    print(f"Date Range: April 25, 2026 - May 24, 2026")
    print(f"Total Shift Records: {len(df_output)}")
    print(f"Unique Clients: {df_output['Client Name'].nunique()}")
    print(f"Total Hours Worked: {df_output['Hours Worked'].sum():,.2f} hours")
    print(f"Total Seconds Worked: {df_output['Seconds Worked'].sum():,.0f} seconds")
    
if __name__ == "__main__":
    main()

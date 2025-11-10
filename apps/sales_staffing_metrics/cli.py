from __future__ import annotations

import argparse
from pathlib import Path

import pandas as pd

from openpyxl import load_workbook

from .views import (
    METRICS_EXPORT_PATH,
    _ensure_headers,
    _find_or_create_week_row,
    _set_cell,
    _load_metrics_export,
    get_workbook_path,
)


def _clear_sheet_data(workbook, sheet_name: str) -> None:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f'Missing "{sheet_name}" sheet in the Sales and Staffing workbook.')

    sheet = workbook[sheet_name]
    if sheet.max_row <= 1:
        return

    sheet.delete_rows(2, sheet.max_row - 1)


def rebuild_from_export(metrics_path: Path, workbook_path: Path) -> None:
    metrics_df = _load_metrics_export(metrics_path)
    if metrics_df.empty:
        raise ValueError(
            f"No metrics available in export '{metrics_path}'. "
            "Upload payroll data at least once to generate an export."
        )

    workbook = load_workbook(filename=workbook_path)

    _clear_sheet_data(workbook, "Revenue")
    _clear_sheet_data(workbook, "Shift Count")

    revenue_sheet = workbook["Revenue"]
    shift_sheet = workbook["Shift Count"]

    revenue_headers = _ensure_headers(revenue_sheet)
    shift_headers = _ensure_headers(shift_sheet)

    metrics_df = metrics_df.dropna(subset=["week_ending"])
    metrics_df = metrics_df.sort_values("week_ending")

    for _, row in metrics_df.iterrows():
        week_ending = pd.to_datetime(row["week_ending"])

        revenue_row, revenue_headers = _find_or_create_week_row(
            revenue_sheet, revenue_headers, week_ending
        )
        shift_row, shift_headers = _find_or_create_week_row(
            shift_sheet, shift_headers, week_ending
        )

        _set_cell(revenue_sheet, revenue_row, revenue_headers, "2025 Revenue", row["total_revenue"])
        _set_cell(
            revenue_sheet,
            revenue_row,
            revenue_headers,
            "New Sales Revenue",
            row.get("new_sales_revenue", 0.0),
        )
        _set_cell(
            revenue_sheet,
            revenue_row,
            revenue_headers,
            "New Sales % of Revenue",
            row.get("new_sales_pct", 0.0),
        )

        _set_cell(shift_sheet, shift_row, shift_headers, "2025 (Shift Count)", row["shift_count"])
        _set_cell(
            shift_sheet,
            shift_row,
            shift_headers,
            "2025 (Fill Rate)",
            row.get("fill_rate", 0.0),
        )

    workbook.save(workbook_path)


def main(argv: list[str] | None = None) -> None:
    parser = argparse.ArgumentParser(
        description="Rebuild the Sales and Staffing workbook from the metrics export."
    )
    default_workbook_path = get_workbook_path()
    parser.add_argument(
        "--metrics",
        type=Path,
        default=None,
        help=f"Path to the CSV export (default: {METRICS_EXPORT_PATH})",
    )
    parser.add_argument(
        "--workbook",
        type=Path,
        default=None,
        help=f"Path to the Excel workbook template (default: {default_workbook_path})",
    )

    args = parser.parse_args(argv)

    metrics_path = args.metrics or METRICS_EXPORT_PATH
    workbook_path = args.workbook or default_workbook_path

    rebuild_from_export(metrics_path=metrics_path, workbook_path=workbook_path)


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main()

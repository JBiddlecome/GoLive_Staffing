from __future__ import annotations

import io
import json
import logging
import os
from datetime import datetime, timedelta, date
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from dateutil.relativedelta import relativedelta
from fastapi import APIRouter, Body, Form, HTTPException, Request, UploadFile, File
from fastapi.concurrency import run_in_threadpool
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse
from openpyxl import load_workbook

templates = Jinja2Templates(directory="templates")
router = APIRouter()

_MODULE_BASE_DIR = Path(__file__).resolve().parents[2]

app_root_env = os.getenv("APP_ROOT")
if app_root_env:
    base_dir_candidate = Path(app_root_env)
    if not base_dir_candidate.is_absolute():
        base_dir_candidate = _MODULE_BASE_DIR / base_dir_candidate
    BASE_DIR = base_dir_candidate
else:
    BASE_DIR = _MODULE_BASE_DIR

DATA_DIR = BASE_DIR / "data"
WORKBOOK_FILENAME = "Sales and Staffing Charts.xlsx"

logger = logging.getLogger("apps.sales_staffing_metrics")
if not logger.handlers:
    handler = logging.StreamHandler()
    formatter = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(name)s: %(message)s"
    )
    handler.setFormatter(formatter)
    logger.addHandler(handler)
logger.setLevel(logging.INFO)


def _resolve_workbook_path() -> Path:
    """
    Resolve the Sales & Staffing workbook path.

    Search order:
      1) SALES_STAFFING_WORKBOOK env override (absolute or relative to BASE_DIR)
      2) data/ in the working directory
      3) repo root (BASE_DIR)
      4) anywhere under BASE_DIR (recursive)
    """

    logger.debug(
        "Resolving workbook path. BASE_DIR=%s DATA_DIR=%s", BASE_DIR, DATA_DIR
    )

    env_override = os.getenv("SALES_STAFFING_WORKBOOK")
    if env_override:
        override_path = Path(env_override)
        if not override_path.is_absolute():
            override_path = BASE_DIR / override_path
        logger.info(
            "SALES_STAFFING_WORKBOOK override provided. Raw value=%s resolved=%s",
            env_override,
            override_path,
        )
        if override_path.exists():
            logger.info("Using workbook override path: %s", override_path)
            return override_path
        logger.warning(
            "Workbook override path does not exist: %s", override_path
        )

    for candidate in (DATA_DIR / WORKBOOK_FILENAME, BASE_DIR / WORKBOOK_FILENAME):
        logger.debug(
            "Checking candidate workbook path: %s (exists=%s)",
            candidate,
            candidate.exists(),
        )
        if candidate.exists():
            logger.info("Found workbook at candidate path: %s", candidate)
            return candidate

    if BASE_DIR.exists():
        logger.debug("Searching recursively under %s for workbook", BASE_DIR)
        for path in BASE_DIR.rglob("*"):
            try:
                if path.name.lower() == WORKBOOK_FILENAME.lower():
                    logger.info("Found workbook via recursive search: %s", path)
                    return path
            except Exception:  # pragma: no cover - defensive
                continue

    # fall back to preferred location (even if missing)
    fallback_path = DATA_DIR / WORKBOOK_FILENAME
    logger.warning(
        "Workbook not found. Falling back to expected location: %s", fallback_path
    )
    return fallback_path


WORKBOOK_PATH = _resolve_workbook_path()
METRICS_EXPORT_PATH = DATA_DIR / "sales_staffing_metrics.csv"
DASHBOARD_DATA_PATH = DATA_DIR / "sales_staffing_dashboard.json"
DEALS_DATA_PATH = DATA_DIR / "sales_staffing_deals.json"
PAYROLL_CANDIDATE_FILENAMES = [
    "payroll 2.csv",
    "Payroll 2.csv",
]


def _resolve_payroll_source_path() -> Path:
    """Return the path to the payroll CSV, tolerating filename casing."""

    search_roots = [DATA_DIR, BASE_DIR]

    for root in search_roots:
        for filename in PAYROLL_CANDIDATE_FILENAMES:
            candidate = root / filename
            if candidate.exists():
                return candidate

    for root in search_roots:
        for candidate in root.glob("*.csv"):
            if candidate.name.lower() == "payroll 2.csv":
                return candidate

    # Fall back to the preferred lowercase filename so callers get a sensible path
    return DATA_DIR / PAYROLL_CANDIDATE_FILENAMES[0]


def _normalize_week_ending(value: datetime) -> datetime:
    """Return the naive midnight datetime used for exports."""

    if value.tzinfo is not None:
        value = value.astimezone(tz=None)
    return value.replace(hour=0, minute=0, second=0, microsecond=0)


def _load_metrics_export(path: Path = METRICS_EXPORT_PATH) -> pd.DataFrame:
    if not path.exists():
        return pd.DataFrame(
            columns=[
                "week_ending",
                "total_revenue",
                "new_sales_revenue",
                "new_sales_pct",
                "shift_count",
                "open_shifts",
                "fill_rate",
            ]
        )

    df = pd.read_csv(path)
    if "week_ending" in df.columns:
        df["week_ending"] = pd.to_datetime(df["week_ending"], errors="coerce")
    return df


def _write_metrics_export(metrics: Dict[str, Any], path: Path = METRICS_EXPORT_PATH) -> None:
    metrics = metrics.copy()
    metrics["week_ending"] = _normalize_week_ending(metrics["week_ending"])

    path.parent.mkdir(parents=True, exist_ok=True)
    export_df = _load_metrics_export(path)
    if not export_df.empty:
        export_df = export_df[export_df["week_ending"] != metrics["week_ending"]]

    export_df = pd.concat([export_df, pd.DataFrame([metrics])], ignore_index=True)
    export_df = export_df.sort_values("week_ending")
    export_df.to_csv(path, index=False)


def _load_dashboard_data(path: Path = DASHBOARD_DATA_PATH) -> Dict[str, Any]:
    if not path.exists():
        return {}

    try:
        with path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:  # pragma: no cover - defensive
        return {}


def _parse_deal_date(value: Any) -> date | None:
    if not value:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, str):
        try:
            return datetime.strptime(value.strip(), "%Y-%m-%d").date()
        except ValueError:
            return None
    return None


def _normalize_deal_entry(entry: Dict[str, Any], *, strict: bool = False) -> Dict[str, str] | None:
    if not isinstance(entry, dict):
        return None

    client = str(entry.get("clientName") or entry.get("client") or "").strip()
    location = str(entry.get("location") or "").strip()
    parsed_date = _parse_deal_date(entry.get("date"))

    if strict:
        if not client:
            raise ValueError("Client name is required.")
        if not location:
            raise ValueError("Location is required.")
        if parsed_date is None:
            raise ValueError("Date must be provided in YYYY-MM-DD format.")

    if not client or not location or parsed_date is None:
        return None

    return {
        "clientName": client,
        "location": location,
        "date": parsed_date.isoformat(),
    }


def _deal_sort_key(entry: Dict[str, Any]) -> Tuple[int, date]:
    entry_date = _parse_deal_date(entry.get("date"))
    if entry_date is None:
        return (10_000, date.max)

    delta_days = abs((entry_date - date.today()).days)
    return (delta_days, entry_date)


def _sort_deal_entries(entries: List[Dict[str, str]]) -> List[Dict[str, str]]:
    return sorted(entries, key=_deal_sort_key)


def _load_deal_tables(path: Path = DEALS_DATA_PATH) -> Dict[str, List[Dict[str, str]]]:
    if not path.exists():
        return {"closed": [], "upcoming": []}

    try:
        with path.open("r", encoding="utf-8") as fh:
            data = json.load(fh)
    except Exception:  # pragma: no cover - defensive
        return {"closed": [], "upcoming": []}

    result: Dict[str, List[Dict[str, str]]] = {"closed": [], "upcoming": []}
    for key in ("closed", "upcoming"):
        entries = data.get(key, []) if isinstance(data, dict) else []
        normalized: List[Dict[str, str]] = []
        if isinstance(entries, list):
            for entry in entries:
                normalized_entry = _normalize_deal_entry(entry)
                if normalized_entry is not None:
                    normalized.append(normalized_entry)
        result[key] = _sort_deal_entries(normalized)
    return result


def _write_deal_tables(
    tables: Dict[str, List[Dict[str, Any]]], path: Path = DEALS_DATA_PATH
) -> Dict[str, List[Dict[str, str]]]:
    cleaned: Dict[str, List[Dict[str, str]]] = {"closed": [], "upcoming": []}

    for key in cleaned:
        entries = tables.get(key, []) if isinstance(tables, dict) else []
        normalized: List[Dict[str, str]] = []
        if isinstance(entries, list):
            for entry in entries:
                normalized_entry = _normalize_deal_entry(entry, strict=True)
                normalized.append(normalized_entry)
        cleaned[key] = _sort_deal_entries(normalized)

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(cleaned, fh, indent=2, ensure_ascii=False)

    return cleaned


def _load_payroll_csv(path: Path | None = None) -> pd.DataFrame:
    if path is None:
        path = _resolve_payroll_source_path()

    if not path.exists():
        return pd.DataFrame()

    try:
        df = pd.read_csv(path, encoding="utf-8-sig", low_memory=False)
    except Exception:  # pragma: no cover - defensive
        return pd.DataFrame()

    if "Date" in df.columns:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce")
    if "Client Won Date" in df.columns:
        df["Client Won Date"] = pd.to_datetime(df["Client Won Date"], errors="coerce")
    return df


def _write_dashboard_data(data: Dict[str, Any], path: Path = DASHBOARD_DATA_PATH) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    safe_data = data.copy()
    if "weekEnding" in safe_data and hasattr(safe_data["weekEnding"], "isoformat"):
        safe_data["weekEnding"] = safe_data["weekEnding"].isoformat()
    if "weekLabel" in safe_data and hasattr(safe_data["weekLabel"], "strftime"):
        safe_data["weekLabel"] = safe_data["weekLabel"].strftime("%B %d, %Y")
    if "newClients" in safe_data:
        serialized_clients: List[Dict[str, Any]] = []
        for client in safe_data["newClients"]:
            client_copy = client.copy()
            won_date_value = client_copy.get("wonDate")
            if hasattr(won_date_value, "isoformat"):
                client_copy["wonDate"] = won_date_value.isoformat()
            won_label_value = client_copy.get("wonDateLabel")
            if hasattr(won_label_value, "strftime"):
                client_copy["wonDateLabel"] = won_label_value.strftime("%B %d, %Y")
            serialized_clients.append(client_copy)
        safe_data["newClients"] = serialized_clients
    if "industries" in safe_data:
        serialized_industries: List[Dict[str, Any]] = []
        for industry in safe_data["industries"]:
            serialized_industries.append(industry.copy())
        safe_data["industries"] = serialized_industries
    with path.open("w", encoding="utf-8") as fh:
        json.dump(safe_data, fh, indent=2)


def _empty_chart_payload() -> Dict[str, Any]:
    return {"weeks": [], "selectedWeek": None}


def _to_date_series(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce")


def _next_sunday_on_or_after(value: datetime) -> datetime:
    days_ahead = (6 - value.weekday()) % 7
    return value + timedelta(days=days_ahead)


def _normalize_money(value: Any) -> float:
    if pd.isna(value):
        return 0.0
    if isinstance(value, str):
        value = value.replace(",", "").replace("$", "").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _normalize_money_nullable(value: Any) -> float | None:
    if pd.isna(value):
        return None
    if isinstance(value, str):
        value = value.replace(",", "").replace("$", "").strip()
        if not value:
            return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _ensure_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for idx, cell in enumerate(ws[1], start=1):
        if cell.value is None:
            continue
        headers[str(cell.value).strip()] = idx
    return headers


def _find_or_create_week_row(
    ws, headers: Dict[str, int], week_ending: datetime
) -> Tuple[int, Dict[str, int]]:
    week_column_name = None
    candidate_week_columns = (
        "Week Ending",
        "Week Ending (Shift Count)",
        "Week Ending (Fill Rate)",
    )

    for candidate in candidate_week_columns:
        if candidate in headers:
            week_column_name = candidate
            break

    if week_column_name is None and "2025 (Shift Count)" in headers:
        insert_position = headers["2025 (Shift Count)"]
        ws.insert_cols(insert_position)
        ws.cell(row=1, column=insert_position).value = "Week Ending"
        headers = _ensure_headers(ws)
        week_column_name = "Week Ending"

    if week_column_name is None:
        raise ValueError(
            f'Required column "Week Ending" not found in sheet "{ws.title}". '
            "Expected one of: Week Ending, Week Ending (Shift Count), Week Ending (Fill Rate), "
            "or 2025 (Shift Count)."
        )

    column = headers[week_column_name]
    for row in range(2, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=column).value
        if cell_value is None:
            continue
        try:
            cell_date = pd.to_datetime(cell_value).date()
        except Exception:  # pragma: no cover - defensive
            continue
        if cell_date == week_ending.date():
            return row, headers

    row = ws.max_row + 1 if ws.max_row >= 1 else 2
    ws.cell(row=row, column=column).value = week_ending
    return row, headers


def _set_cell(ws, row: int, headers: Dict[str, int], column: str, value: Any) -> None:
    if column not in headers:
        raise ValueError(f'Required column "{column}" not found in sheet "{ws.title}"')
    ws.cell(row=row, column=headers[column]).value = value


def _clean_int(value: Any) -> int | None:
    if pd.isna(value):
        return None
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return None


def _clean_float(value: Any) -> float | None:
    if pd.isna(value):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _normalize_excel_header(header: Any) -> str:
    """Return a normalized representation of an Excel header cell."""

    if header is None:
        return ""

    text = str(header)
    # Collapse consecutive whitespace and strip leading/trailing spaces so headers such as
    # "Week Ending (Shift Count) " or "Week  Ending (Shift Count)" still match.
    normalized = " ".join(text.split()).strip().lower()
    return normalized


def _load_chart_data(path: Path | None = None) -> Dict[str, Any]:
    if path is None:
        path = _resolve_workbook_path()

    logger.debug("Loading chart data from workbook path: %s", path)

    if not path.exists():
        logger.warning("Workbook path does not exist when loading chart data: %s", path)
        return _empty_chart_payload()

    try:
        with pd.ExcelFile(path) as workbook:
            target_sheet = next(
                (
                    sheet_name
                    for sheet_name in workbook.sheet_names
                    if _normalize_excel_header(sheet_name) == "shift count"
                ),
                None,
            )

            if target_sheet is None:
                logger.warning(
                    "Shift Count sheet not found in workbook '%s'. Sheets available: %s",
                    path,
                    workbook.sheet_names,
                )
                return _empty_chart_payload()

            shift_df = workbook.parse(target_sheet)
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception(
            "Failed to load chart data from workbook '%s': %s", path, exc
        )
        return _empty_chart_payload()

    if shift_df.empty:
        logger.warning(
            "Shift Count sheet '%s' in workbook '%s' is empty", target_sheet, path
        )
        return _empty_chart_payload()

    shift_df = shift_df.copy()
    column_lookup = {
        _normalize_excel_header(column): column for column in shift_df.columns if column is not None
    }

    week_columns: List[str] = [
        column_lookup[normalized]
        for normalized in (
            "week ending (shift count)",
            "week ending (fill rate)",
            "week ending",
        )
        if normalized in column_lookup
    ]

    if not week_columns:
        logger.warning(
            "No week ending columns found in sheet '%s' of workbook '%s'. Headers: %s",
            target_sheet,
            path,
            list(shift_df.columns),
        )
        return _empty_chart_payload()

    shift_2024_column = column_lookup.get("2024 (shift count)")
    shift_2025_column = column_lookup.get("2025 (shift count)")
    fill_2024_column = column_lookup.get("2024 (fill rate)")
    fill_2025_column = column_lookup.get("2025 (fill rate)")

    records: List[Tuple[date, Dict[str, Any]]] = []
    for _, row in shift_df.iterrows():
        week_ts = None
        for column in week_columns:
            value = row.get(column)
            if pd.isna(value):
                continue
            try:
                parsed = pd.to_datetime(value, errors="coerce")
            except Exception:  # pragma: no cover - defensive
                parsed = pd.NaT
            if pd.isna(parsed):
                continue
            week_ts = parsed.to_pydatetime()
            break

        if week_ts is None:
            continue

        week_date = week_ts.date()
        record = {
            "weekEnding": week_ts.strftime("%Y-%m-%d"),
            "label": week_ts.strftime("%B %d, %Y"),
            "shiftCount2024": _clean_int(row.get(shift_2024_column)) if shift_2024_column else None,
            "shiftCount2025": _clean_int(row.get(shift_2025_column)) if shift_2025_column else None,
            "fillRate2024": _clean_float(row.get(fill_2024_column)) if fill_2024_column else None,
            "fillRate2025": _clean_float(row.get(fill_2025_column)) if fill_2025_column else None,
        }
        records.append((week_date, record))

    if not records:
        logger.info(
            "No records with valid week ending values found in workbook '%s'", path
        )
        return _empty_chart_payload()

    records.sort(key=lambda item: item[0])
    weeks = [record for _, record in records]

    payload = {"weeks": weeks, "selectedWeek": weeks[-1]["weekEnding"]}
    logger.debug(
        "Loaded %s weeks of chart data from '%s'. Selected week=%s",
        len(weeks),
        path,
        payload["selectedWeek"],
    )
    return payload


def _load_revenue_goal_data(path: Path | None = None) -> List[Dict[str, Any]]:
    if path is None:
        path = _resolve_workbook_path()

    if not path.exists():
        logger.warning(
            "Workbook path does not exist when loading revenue data: %s", path
        )
        return []

    try:
        with pd.ExcelFile(path) as workbook:
            target_sheet = next(
                (
                    sheet_name
                    for sheet_name in workbook.sheet_names
                    if _normalize_excel_header(sheet_name) == "revenue"
                ),
                None,
            )

            if target_sheet is None:
                logger.warning(
                    "Revenue sheet not found in workbook '%s'. Sheets available: %s",
                    path,
                    workbook.sheet_names,
                )
                return []

            revenue_df = workbook.parse(target_sheet)
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception(
            "Failed to load revenue data from workbook '%s': %s", path, exc
        )
        return []

    if revenue_df.empty:
        logger.warning(
            "Revenue sheet '%s' in workbook '%s' is empty", target_sheet, path
        )
        return []

    revenue_df = revenue_df.copy()
    column_lookup = {
        _normalize_excel_header(column): column
        for column in revenue_df.columns
        if column is not None
    }

    week_column = column_lookup.get("week ending")
    revenue_column = column_lookup.get("2025 revenue")
    goal_column = column_lookup.get("2025 revenue goal")
    new_sales_revenue_column = column_lookup.get("new sales revenue")
    new_sales_pct_column = column_lookup.get("new sales % of revenue")

    if not week_column or not revenue_column or not goal_column:
        logger.warning(
            "Required columns missing from revenue sheet '%s'. Headers: %s",
            target_sheet,
            list(revenue_df.columns),
        )
        return []

    records: List[Tuple[date, Dict[str, Any]]] = []
    for _, row in revenue_df.iterrows():
        week_value = row.get(week_column)
        if pd.isna(week_value):
            continue
        week_ts = pd.to_datetime(week_value, errors="coerce")
        if pd.isna(week_ts):
            continue

        revenue_value = _clean_float(row.get(revenue_column))
        goal_value = _clean_float(row.get(goal_column))

        week_dt = week_ts.to_pydatetime().date()
        new_sales_revenue_value = (
            _clean_float(row.get(new_sales_revenue_column))
            if new_sales_revenue_column
            else None
        )
        new_sales_pct_value = (
            _clean_float(row.get(new_sales_pct_column))
            if new_sales_pct_column
            else None
        )

        records.append(
            (
                week_dt,
                {
                    "weekEnding": week_dt.strftime("%Y-%m-%d"),
                    "label": week_ts.strftime("%B %d, %Y"),
                    "revenue2025": revenue_value,
                    "revenueGoal2025": goal_value,
                    "newSalesRevenue": new_sales_revenue_value,
                    "newSalesPct": new_sales_pct_value,
                },
            )
        )

    if not records:
        return []

    records.sort(key=lambda item: item[0])
    return [record for _, record in records]


def _summarize_top_clients(df: pd.DataFrame) -> List[Dict[str, Any]]:
    if df.empty or "Client" not in df.columns or "Total Bill" not in df.columns:
        return []

    working = df.copy()
    working["Client"] = working["Client"].fillna("Unknown Client").astype(str)
    working["Total Bill"] = working["Total Bill"].apply(_normalize_money)

    if "Bill Rate" in working.columns:
        working["Bill Rate"] = working["Bill Rate"].apply(_normalize_money_nullable)
    else:
        working["Bill Rate"] = pd.NA

    working["Bill Rate"] = pd.to_numeric(working["Bill Rate"], errors="coerce")

    grouped = (
        working.groupby("Client", as_index=False)
        .agg(total_bill=("Total Bill", "sum"), average_bill_rate=("Bill Rate", "mean"))
        .sort_values("total_bill", ascending=False)
        .head(5)
    )

    results: List[Dict[str, Any]] = []
    for _, row in grouped.iterrows():
        total_bill_value = float(row["total_bill"]) if pd.notna(row["total_bill"]) else 0.0
        avg_bill_rate_value = (
            float(row["average_bill_rate"]) if pd.notna(row["average_bill_rate"]) else None
        )
        results.append(
            {
                "client": row["Client"],
                "totalBill": total_bill_value,
                "revenue": total_bill_value,
                "averageBillRate": avg_bill_rate_value,
            }
        )

    return results


def _calculate_top_clients(payroll_df: pd.DataFrame) -> List[Dict[str, Any]]:
    return _summarize_top_clients(payroll_df)


def _calculate_top_clients_by_week(
    payroll_df: pd.DataFrame, week_endings: List[str]
) -> Dict[str, List[Dict[str, Any]]]:
    results: Dict[str, List[Dict[str, Any]]] = {week: [] for week in week_endings}

    if payroll_df.empty or "Date" not in payroll_df.columns:
        return results

    working = payroll_df.copy()
    working["Date"] = pd.to_datetime(working["Date"], errors="coerce")
    working = working.dropna(subset=["Date"])

    if working.empty:
        return results

    working["Date"] = working["Date"].dt.normalize()

    for week in week_endings:
        try:
            week_date = datetime.strptime(week, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            results[week] = []
            continue

        start_date = week_date - timedelta(days=6)
        mask = (working["Date"].dt.date >= start_date) & (
            working["Date"].dt.date <= week_date
        )
        subset = working.loc[mask]
        results[week] = _summarize_top_clients(subset)

    return results


def _calculate_new_clients_by_week(
    payroll_df: pd.DataFrame, week_endings: List[str]
) -> Dict[str, List[Dict[str, Any]]]:
    results: Dict[str, List[Dict[str, Any]]] = {week: [] for week in week_endings}

    required_columns = {"Date", "Client", "Client Won Date", "Total Bill"}
    if payroll_df.empty or not required_columns.issubset(payroll_df.columns):
        return results

    df = payroll_df.copy()
    df["Date"] = pd.to_datetime(df["Date"], errors="coerce").dt.normalize()
    df["Client Won Date"] = pd.to_datetime(
        df["Client Won Date"], errors="coerce"
    ).dt.normalize()
    df = df.dropna(subset=["Date", "Client Won Date"])
    if df.empty:
        return results

    df["Client"] = df["Client"].fillna("Unknown Client").astype(str)
    df["Total Bill"] = df["Total Bill"].apply(_normalize_money)

    for week in week_endings:
        try:
            week_date = datetime.strptime(week, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            results[week] = []
            continue

        start_of_week = week_date - timedelta(days=6)
        six_months_prior = week_date - relativedelta(months=6)

        weekly_subset = df[
            (df["Date"].dt.date >= start_of_week)
            & (df["Date"].dt.date <= week_date)
            & (df["Client Won Date"].dt.date >= six_months_prior)
            & (df["Client Won Date"].dt.date <= week_date)
        ]

        if weekly_subset.empty:
            results[week] = []
            continue

        grouped = (
            weekly_subset.groupby("Client", as_index=False)
            .agg(
                total_bill=("Total Bill", "sum"),
                won_date=("Client Won Date", "max"),
            )
            .sort_values("won_date", ascending=False)
        )

        entries: List[Dict[str, Any]] = []
        for _, row in grouped.iterrows():
            total_bill_value = (
                float(row["total_bill"]) if pd.notna(row["total_bill"]) else 0.0
            )
            won_iso, won_label = _format_won_date(row["won_date"])
            won_dt = pd.to_datetime(row["won_date"], errors="coerce")
            is_highlighted = False
            if pd.notna(won_dt):
                won_date_value = won_dt.date()
                is_highlighted = start_of_week <= won_date_value <= week_date

            entries.append(
                {
                    "client": row["Client"],
                    "totalBill": total_bill_value,
                    "wonDate": won_iso,
                    "wonDateLabel": won_label,
                    "isHighlighted": is_highlighted,
                }
            )

        results[week] = entries

    return results


def _format_won_date(value: Any) -> Tuple[str | None, str | None]:
    if pd.isna(value):
        return None, None
    if isinstance(value, pd.Timestamp):
        won_date = value.to_pydatetime()
    elif isinstance(value, datetime):
        won_date = value
    else:
        try:
            won_date = pd.to_datetime(value)
            if pd.isna(won_date):
                return None, None
            won_date = won_date.to_pydatetime()
        except Exception:  # pragma: no cover - defensive
            return None, None
    return won_date.date().isoformat(), won_date.strftime("%B %d, %Y")


def _calculate_new_clients(
    payroll_df: pd.DataFrame,
    six_months_prior: datetime,
    week_ending: datetime,
    highlight_range: Tuple[datetime, datetime] | None = None,
    copy_frame: bool = True,
) -> List[Dict[str, Any]]:
    required_columns = {"Client", "Client Won Date", "Total Bill"}
    if not required_columns.issubset(payroll_df.columns):
        return []

    valid_won_mask = payroll_df["Client Won Date"].notna()
    if not valid_won_mask.any():
        return []

    df = payroll_df.loc[valid_won_mask]
    if copy_frame:
        df = df.copy()

    df["Client"] = df["Client"].fillna("Unknown Client").astype(str)
    if "Total Bill" in df.columns and df["Total Bill"].dtype == "O":
        df["Total Bill"] = df["Total Bill"].apply(_normalize_money)
    if "Client Won Date" in df.columns and not pd.api.types.is_datetime64_any_dtype(
        df["Client Won Date"]
    ):
        df["Client Won Date"] = pd.to_datetime(df["Client Won Date"], errors="coerce")
        df = df[df["Client Won Date"].notna()].copy()
        if df.empty:
            return []

    mask = (df["Client Won Date"] >= six_months_prior) & (
        df["Client Won Date"] <= week_ending
    )
    df = df.loc[mask]
    if df.empty:
        return []

    grouped = (
        df.groupby("Client", as_index=False)
        .agg(total_bill=("Total Bill", "sum"), won_date=("Client Won Date", "max"))
        .sort_values("won_date", ascending=False)
    )

    if highlight_range is not None:
        highlight_start, highlight_end = highlight_range
        highlight_start_ts = pd.Timestamp(highlight_start)
        highlight_end_ts = pd.Timestamp(highlight_end)
        grouped["is_highlighted"] = grouped["won_date"].apply(
            lambda value: pd.notna(value)
            and highlight_start_ts <= pd.Timestamp(value) <= highlight_end_ts
        )
    else:
        grouped["is_highlighted"] = False

    results: List[Dict[str, Any]] = []
    for _, row in grouped.iterrows():
        total_bill_value = float(row["total_bill"]) if pd.notna(row["total_bill"]) else 0.0
        won_iso, won_label = _format_won_date(row["won_date"])
        results.append(
            {
                "client": row["Client"],
                "totalBill": total_bill_value,
                "wonDate": won_iso,
                "wonDateLabel": won_label,
                "isHighlighted": bool(row.get("is_highlighted", False)),
            }
        )

    return results


def _summarize_industry_totals(df: pd.DataFrame) -> List[Dict[str, Any]]:
    if df.empty or "Industry" not in df.columns or "Total Bill" not in df.columns:
        return []

    working = df.copy()
    working["Industry"] = working["Industry"].fillna("Unknown Industry").astype(str)
    working["Total Bill"] = working["Total Bill"].apply(_normalize_money)

    grouped = (
        working.groupby("Industry", as_index=False)
        .agg(total_bill=("Total Bill", "sum"))
        .sort_values("total_bill", ascending=False)
    )

    results: List[Dict[str, Any]] = []
    for _, row in grouped.iterrows():
        total_bill_value = float(row["total_bill"]) if pd.notna(row["total_bill"]) else 0.0
        results.append({"industry": row["Industry"], "totalBill": total_bill_value})

    return results


def _calculate_industry_totals(payroll_df: pd.DataFrame) -> List[Dict[str, Any]]:
    return _summarize_industry_totals(payroll_df)


def _calculate_industry_totals_by_week(
    payroll_df: pd.DataFrame, week_endings: List[str]
) -> Dict[str, List[Dict[str, Any]]]:
    results: Dict[str, List[Dict[str, Any]]] = {week: [] for week in week_endings}

    required_columns = {"Date", "Industry", "Total Bill"}
    if payroll_df.empty or not required_columns.issubset(payroll_df.columns):
        return results

    working = payroll_df.copy()
    working["Date"] = pd.to_datetime(working["Date"], errors="coerce").dt.normalize()
    working = working.dropna(subset=["Date"])

    if working.empty:
        return results

    working["Industry"] = working["Industry"].fillna("Unknown Industry").astype(str)
    working["Total Bill"] = working["Total Bill"].apply(_normalize_money)

    for week in week_endings:
        try:
            week_date = datetime.strptime(week, "%Y-%m-%d").date()
        except (TypeError, ValueError):
            results[week] = []
            continue

        start_date = week_date - timedelta(days=6)
        mask = (working["Date"].dt.date >= start_date) & (
            working["Date"].dt.date <= week_date
        )
        weekly_subset = working.loc[mask]
        results[week] = _summarize_industry_totals(weekly_subset)

    return results


def _update_workbook(payroll_df: pd.DataFrame, open_shifts: int) -> Dict[str, Any]:
    workbook_path = _resolve_workbook_path()
    if not workbook_path.exists():
        raise FileNotFoundError(
            "Sales and Staffing workbook not found. Expected to find it at "
            f"{workbook_path}."
        )

    payroll_df = payroll_df.copy()
    if "Total Bill" not in payroll_df.columns:
        raise ValueError('"Total Bill" column not found in the payroll workbook.')
    if "Date" not in payroll_df.columns:
        raise ValueError('"Date" column not found in the payroll workbook.')

    payroll_df["Date"] = _to_date_series(payroll_df["Date"])
    if payroll_df["Date"].dropna().empty:
        raise ValueError("No valid dates found in the payroll workbook.")

    if "Client Won Date" in payroll_df.columns:
        payroll_df["Client Won Date"] = _to_date_series(payroll_df["Client Won Date"])
    else:
        payroll_df["Client Won Date"] = pd.NaT

    max_date = payroll_df["Date"].max()
    week_ending = _next_sunday_on_or_after(max_date)

    payroll_df["Total Bill"] = payroll_df["Total Bill"].apply(_normalize_money)
    total_revenue = float(payroll_df["Total Bill"].sum())

    six_months_prior = week_ending - relativedelta(months=6)
    mask_new_sales = (
        payroll_df["Client Won Date"].notna()
        & (payroll_df["Client Won Date"] >= six_months_prior)
        & (payroll_df["Client Won Date"] <= week_ending)
    )
    new_sales_revenue = float(payroll_df.loc[mask_new_sales, "Total Bill"].sum())
    new_sales_pct = new_sales_revenue / total_revenue if total_revenue > 0 else 0.0

    shift_count = int((payroll_df["Total Bill"] > 0).sum())
    open_shifts = max(0, int(open_shifts))
    total_shifts = shift_count + open_shifts
    fill_rate = shift_count / total_shifts if total_shifts > 0 else 0.0

    workbook = load_workbook(filename=workbook_path)

    if "Revenue" not in workbook.sheetnames:
        raise ValueError('Missing "Revenue" sheet in the Sales and Staffing workbook.')
    revenue_sheet = workbook["Revenue"]
    revenue_headers = _ensure_headers(revenue_sheet)
    revenue_row, revenue_headers = _find_or_create_week_row(
        revenue_sheet, revenue_headers, week_ending
    )
    _set_cell(revenue_sheet, revenue_row, revenue_headers, "2025 Revenue", total_revenue)
    _set_cell(revenue_sheet, revenue_row, revenue_headers, "New Sales Revenue", new_sales_revenue)
    _set_cell(revenue_sheet, revenue_row, revenue_headers, "New Sales % of Revenue", new_sales_pct)

    if "Shift Count" not in workbook.sheetnames:
        raise ValueError('Missing "Shift Count" sheet in the Sales and Staffing workbook.')
    shift_sheet = workbook["Shift Count"]
    shift_headers = _ensure_headers(shift_sheet)
    shift_row, shift_headers = _find_or_create_week_row(
        shift_sheet, shift_headers, week_ending
    )
    _set_cell(shift_sheet, shift_row, shift_headers, "2025 (Shift Count)", shift_count)
    _set_cell(shift_sheet, shift_row, shift_headers, "2025 (Fill Rate)", fill_rate)

    workbook.save(workbook_path)

    metrics = {
        "week_ending": week_ending,
        "total_revenue": total_revenue,
        "new_sales_revenue": new_sales_revenue,
        "new_sales_pct": new_sales_pct,
        "shift_count": shift_count,
        "open_shifts": open_shifts,
        "fill_rate": fill_rate,
    }

    _write_metrics_export(metrics)

    top_clients = _calculate_top_clients(payroll_df)
    highlight_range = (week_ending - timedelta(days=6), week_ending)
    new_clients = _calculate_new_clients(
        payroll_df, six_months_prior, week_ending, highlight_range
    )
    industry_totals = _calculate_industry_totals(payroll_df)
    dashboard_payload = {
        "weekEnding": week_ending,
        "weekLabel": week_ending,
        "topClients": top_clients,
        "newClients": new_clients,
        "industries": industry_totals,
    }
    _write_dashboard_data(dashboard_payload)

    return metrics


def _read_payroll(upload: UploadFile) -> pd.DataFrame:
    try:
        upload.file.seek(0)
        payload = upload.file.read()
        if not payload:
            raise ValueError("The uploaded payroll workbook is empty.")
        logger.debug(
            "Reading payroll workbook '%s' (%s bytes)",
            upload.filename,
            len(payload),
        )
        return pd.read_excel(io.BytesIO(payload), engine="openpyxl")
    except ValueError:
        raise
    except Exception as exc:  # pragma: no cover - defensive
        logger.exception(
            "Unable to read Excel file '%s'", upload.filename,
        )
        raise ValueError(f"Unable to read Excel file '{upload.filename}'.") from exc


def _build_chart_payload() -> Dict[str, Any]:
    chart_payload = _load_chart_data()

    weeks = [
        week.get("weekEnding")
        for week in chart_payload.get("weeks", [])
        if isinstance(week, dict) and week.get("weekEnding")
    ]

    payroll_df = _load_payroll_csv()
    top_clients_by_week = _calculate_top_clients_by_week(payroll_df, weeks)
    new_clients_by_week = _calculate_new_clients_by_week(payroll_df, weeks)
    industry_totals_by_week = _calculate_industry_totals_by_week(payroll_df, weeks)

    chart_payload["topClientsByWeek"] = top_clients_by_week
    chart_payload["newClientsByWeek"] = new_clients_by_week
    chart_payload["industryTotalsByWeek"] = industry_totals_by_week
    chart_payload["revenueSeries"] = _load_revenue_goal_data()

    return chart_payload


def _build_page_context(**extra: Any) -> Dict[str, Any]:
    chart_payload = _build_chart_payload()

    base_context = {
        "workbook_path": _resolve_workbook_path(),
        "metrics_export_path": METRICS_EXPORT_PATH,
        "chart_data": chart_payload,
        "deal_tables": _load_deal_tables(),
    }
    base_context.update(extra)
    workbook_path = base_context["workbook_path"]
    logger.info(
        "Building page context. workbook_path=%s (exists=%s) metrics_export_path=%s (exists=%s) weeks_available=%s",
        workbook_path,
        workbook_path.exists(),
        METRICS_EXPORT_PATH,
        METRICS_EXPORT_PATH.exists(),
        len(chart_payload.get("weeks", []))
        if isinstance(chart_payload, dict)
        else "unknown",
    )
    return base_context


@router.get("")
async def page(request: Request):
    logger.info("Rendering Sales & Staffing Metrics page for %s", request.client)
    context = _build_page_context(request=request)
    return templates.TemplateResponse("apps/sales_staffing_metrics.html", context)


@router.post("/update")
async def update(
    request: Request,
    payroll: UploadFile = File(...),
    open_shifts: str = Form(""),
):
    logger.info(
        "Received Sales & Staffing metrics update request. payroll_filename=%s open_shifts_raw='%s'",
        payroll.filename,
        open_shifts,
    )
    context = _build_page_context(request=request)

    try:
        payroll_df = await run_in_threadpool(_read_payroll, payroll)
    except ValueError as exc:
        logger.warning("Payroll upload invalid: %s", exc)
        context.update({"error": str(exc)})
        return templates.TemplateResponse(
            "apps/sales_staffing_metrics.html",
            context,
            status_code=400,
        )
    finally:
        await payroll.close()

    try:
        open_shifts_value = int(open_shifts.replace(",", "").strip()) if open_shifts.strip() else 0
    except ValueError:
        logger.warning("Invalid open shifts value provided: '%s'", open_shifts)
        context.update({"error": "Open shifts must be a whole number."})
        return templates.TemplateResponse(
            "apps/sales_staffing_metrics.html",
            context,
            status_code=400,
        )

    try:
        result = await run_in_threadpool(_update_workbook, payroll_df, open_shifts_value)
    except FileNotFoundError as exc:
        logger.error("Workbook update failed - file missing: %s", exc)
        context.update({"error": str(exc)})
        return templates.TemplateResponse(
            "apps/sales_staffing_metrics.html",
            context,
            status_code=404,
        )
    except ValueError as exc:
        logger.warning("Workbook update failed due to invalid data: %s", exc)
        context.update({"error": str(exc)})
        return templates.TemplateResponse(
            "apps/sales_staffing_metrics.html",
            context,
            status_code=400,
        )

    logger.info(
        "Workbook update succeeded. Metrics: %s", result
    )
    context.update({"result": result, "chart_data": _build_chart_payload()})
    return templates.TemplateResponse("apps/sales_staffing_metrics.html", context)


@router.get("/deals")
async def get_deal_tables() -> JSONResponse:
    tables = _load_deal_tables()
    return JSONResponse(tables)


@router.post("/deals")
async def save_deal_tables(payload: Dict[str, Any] = Body(...)) -> JSONResponse:
    try:
        cleaned = _write_deal_tables(payload)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc
    return JSONResponse({"status": "ok", **cleaned})
